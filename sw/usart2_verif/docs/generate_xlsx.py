#!/usr/bin/env python3
"""
generate_xlsx.py — Generate USART2 verification XLSX deliverables

Produces three workbooks in docs/:
  1. USART_TestPlan.xlsx      — complete test plan with coverage mapping
  2. usart_errata.xlsx        — spec-vs-model gap tracker
  3. coverage_summary.xlsx    — coverage metrics by category

Prerequisites:
    pip install openpyxl

Run:
    python3 docs/generate_xlsx.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (Font, Alignment, PatternFill, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

DOCS_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Colour palette ──────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="003366")   # dark navy
ROW_EVEN  = PatternFill("solid", fgColor="E8F0FE")   # light blue
ROW_ODD   = PatternFill("solid", fgColor="FFFFFF")   # white
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")   # green
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")   # red
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")   # amber
HDR_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11)
WRAP      = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER    = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN      = Side(style="thin", color="AAAAAA")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def apply_header(ws, headers, col_widths):
    ws.append(headers)
    for col, width in enumerate(col_widths, 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def style_row(ws, row_num, fill):
    for cell in ws[row_num]:
        cell.fill      = fill
        cell.font      = BODY_FONT
        cell.alignment = WRAP
        cell.border    = BORDER


def auto_row_height(ws):
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 50


# ══════════════════════════════════════════════════════════════════════════════
# 1. USART_TestPlan.xlsx
# ══════════════════════════════════════════════════════════════════════════════
TEST_PLAN = [
    # (ID, Category, Name, Objective, Registers, Interrupts/Flags, Expected, CovType, Priority, Status)
    # ── Actual run results: RST-001 PASS; RST-002..005 intentional FAIL (BUG-007) ──
    ("RST-001","Reset","Power-on Reset Values",
     "Read CON/STATUS/BG/FDR before any write; verify spec reset values",
     "CON,STATUS,BG,FDR","None","All registers==0x00000000 at power-on",
     "Register Coverage","P1","Pass"),
    ("RST-002","Reset","HW Reset Clears CON [BUG-007]",
     "Write CON_INIT; assert SWRST(0x14); verify CON returns to 0",
     "CON,SWRST","None","CON==0 after SWRST (intentional FAIL — BUG-007)",
     "Reset Coverage","P1","Fail"),
    ("RST-003","Reset","HW Reset Clears STATUS [BUG-007]",
     "Write TBUF (TBIR set, MIE=0); assert SWRST; verify STATUS==0",
     "STATUS,SWRST","TBIR","STATUS==0 after SWRST (intentional FAIL — BUG-007)",
     "Reset Coverage","P1","Fail"),
    ("RST-004","Reset","HW Reset Clears rbuf_full [BUG-007]",
     "Fill RBUF_B; assert SWRST on B; next byte must generate RIR (not EIR)",
     "RBUF,SWRST","RIR,EIR","RIR fires (rbuf_full cleared) — intentional FAIL BUG-007",
     "Reset Coverage","P1","Fail"),
    ("RST-005","Reset","HW Reset Mid-Transfer [BUG-007]",
     "Write TBUF; immediately assert SWRST; verify STATUS=0 + no byte delivered",
     "STATUS,SWRST","TBIR,RIR","STATUS==0, no stale byte — intentional FAIL BUG-007",
     "Reset Coverage","P1","Fail"),

    ("RW-001","R/W","CON R/W Patterns",
     "Write 6 different CON patterns; verify readback matches for each",
     "CON","None","Readback==Written all 6 patterns","Register Coverage","P1","Pass"),
    ("RW-002","R/W","CON Reserved Bits [31:16]",
     "Write 0xFFFFFFFF to CON; verify CON[31:16] behaviour",
     "CON","None","Reserved bits behaviour logged as finding","Register Coverage","P2","Finding"),
    ("RW-003","R/W","TBUF Write-Only Triggers TBIR",
     "Write TBUF; verify TBIR fires (no read path)",
     "TBUF","TBIR","TBIR in log within 1 WFI","Register Coverage","P1","Pass"),
    ("RW-004","R/W","RBUF Read-Only Returns Byte",
     "Send byte A→B; read RBUF_B; verify correct value",
     "RBUF","RIR","RBUF_B==sent byte","Register Coverage","P1","Pass"),
    ("RW-005","R/W","STATUS W1C Semantics",
     "Write 0 → bits unchanged; write 1 → bit cleared",
     "STATUS","TBIR","Write-0 preserves, write-1 clears","Register Coverage","P1","Pass"),
    ("RW-006","R/W","BG Not Implemented",
     "Write BG (0x0C); verify VP does not crash; log finding",
     "BG","None","No crash; finding BG-not-implemented logged","Register Coverage","P3","Finding"),
    ("RW-007","R/W","FDR Not Implemented",
     "Write FDR (0x10); verify VP does not crash; log finding",
     "FDR","None","No crash; finding FDR-not-implemented logged","Register Coverage","P3","Finding"),

    ("FN-001","Functional","A→B Single Byte 0x55",
     "Send 0x55 A→B; verify TBIR_A, RIR_B, RBUF_B=0x55",
     "TBUF,RBUF,STATUS","TBIR,RIR","TBIR_A+RIR_B in log, RBUF_B=0x55","Functional Coverage","P1","Pass"),
    ("FN-002","Functional","B→A Single Byte 0xAA",
     "Send 0xAA B→A; verify TBIR_B, RIR_A, RBUF_A=0xAA",
     "TBUF,RBUF,STATUS","TBIR,RIR","TBIR_B+RIR_A in log, RBUF_A=0xAA","Functional Coverage","P1","Pass"),
    ("FN-003","Functional","Overrun EIR",
     "Fill RBUF_B; send second byte; verify EIR_B fires (OEN=1)",
     "CON,RBUF,STATUS","RIR,EIR","EIR_B in log after overrun","Error Coverage","P1","Pass"),
    ("FN-004","Functional","Multi-byte Stream A→B",
     "Send [0xDE,0xAD,0xBE,0xEF]; verify each byte received correctly",
     "TBUF,RBUF,STATUS","TBIR,RIR","All 4 bytes received in order","Functional Coverage","P1","Pass"),
    ("FN-005","Functional","TIR TX-Complete",
     "Send byte; verify TBIR_A fires first, TIR_A fires 2µs later",
     "TBUF,STATUS","TBIR,TIR","TBIR before TIR; both in log","Interrupt Coverage","P1","Pass"),

    ("CLK-001","Clock","TBIR Immediate",
     "Verify TBIR STATUS bit set in same b_transport call as TBUF write",
     "TBUF,STATUS","TBIR","TBIR set within 1 quantum","Timing Coverage","P1","Pass"),
    ("CLK-002","Clock","TIR Separate Interrupt",
     "TIR must be a separate log entry (different ISR) from TBIR",
     "STATUS","TBIR,TIR","TIR log index > TBIR log index","Interrupt Coverage","P1","Pass"),
    ("CLK-003","Clock","RIR After TBIR",
     "RIR_B must appear after TBIR_A in event log",
     "STATUS","TBIR,RIR","RIR log index >= TBIR log index","Timing Coverage","P1","Pass"),
    ("CLK-004","Clock","TBIR→TIR Order (4 transfers)",
     "On 4 consecutive transfers, TBIR always precedes TIR",
     "STATUS","TBIR,TIR","mon_check_order() passes all 4","Interrupt Coverage","P2","Pass"),

    ("LB-001","Loopback","A→B 8 Boundary Values",
     "Send 8 boundary bytes via A→B; verify RBUF_B integrity",
     "TBUF,RBUF","TBIR,RIR","All 8 values correct","Data Coverage","P1","Pass"),
    ("LB-002","Loopback","B→A 8 Boundary Values",
     "Send 8 boundary bytes via B→A; verify RBUF_A integrity",
     "TBUF,RBUF","TBIR,RIR","All 8 values correct","Data Coverage","P1","Pass"),
    ("LB-003","Loopback","Round-Trip Echo A→B→A",
     "Send 0xE5 A→B; B reads and re-sends; A receives 0xE5",
     "TBUF,RBUF","TBIR,RIR","Echo byte intact after round trip","Functional Coverage","P2","Pass"),
    ("LB-004","Loopback","CON.LB Bit Storage",
     "Write CON.LB=1; verify stored; log finding (VP no-op)",
     "CON","None","CON.LB stored; finding logged","Register Coverage","P3","Finding"),

    ("ADV-001","Advanced","EIR Suppressed OEN=0",
     "Overrun with OEN=0; verify EIR does NOT fire",
     "CON,STATUS","EIR","EIR absent in log (OEN=0)","Error Coverage","P1","Pass"),
    ("ADV-002","Advanced","EIR Fires OEN=1",
     "Overrun with OEN=1; verify EIR fires",
     "CON,STATUS","EIR","EIR present in log (OEN=1)","Error Coverage","P1","Pass"),
    ("ADV-003","Advanced","Overrun Recovery",
     "After overrun, drain RBUF; next byte generates RIR correctly",
     "RBUF,STATUS","RIR,EIR","RIR fires again after drain","Functional Coverage","P1","Pass"),
    ("ADV-004","Advanced","Bidirectional Simultaneous",
     "Fire A→B and B→A in same window; verify 4 events + correct data",
     "TBUF,RBUF","TBIR,RIR","4 events; both bytes correct","Functional Coverage","P2","Pass"),
    ("ADV-005","Advanced","STATUS Multi-bit Set",
     "Sample STATUS while TBIR+TIR both set simultaneously",
     "STATUS","TBIR,TIR","STATUS shows ≥1 bit set","Register Coverage","P2","Pass"),
    ("ADV-006","Advanced","RBUF Retention",
     "Verify RBUF data persists; second read returns stale (VP documented behaviour)",
     "RBUF","RIR","First read correct; stale re-read logged","Functional Coverage","P3","Finding"),

    ("PERF-001","Performance","32-byte Throughput No Loss",
     "Send 32 bytes A→B; verify none lost",
     "TBUF,RBUF","TBIR,RIR","0 errors","Performance Coverage","P1","Pass"),
    ("PERF-002","Performance","TBIR Latency ≤1 WFI",
     "TBIR must appear within first WFI after TBUF write",
     "TBUF,STATUS","TBIR","TBIR in log after 1 WFI","Timing Coverage","P1","Pass"),
    ("PERF-003","Performance","RIR Ordering 16 Transfers",
     "RIR always after TBIR on 16 consecutive transfers",
     "STATUS","TBIR,RIR","All 16 pass mon_check_order","Timing Coverage","P2","Pass"),
    ("PERF-004","Performance","Interrupt Count 32 Bytes",
     "32 bytes → 32 TBIR_A + 32 RIR_B minimum",
     "STATUS","TBIR,RIR","TBIR count=32 RIR count=32","Interrupt Coverage","P2","Pass"),

    ("STR-001","Stress","100-byte Sequential",
     "Send 100 bytes A→B with walking pattern; 0 errors expected",
     "TBUF,RBUF","TBIR,RIR","errors=0","Stress Coverage","P1","Pass"),
    ("STR-002","Stress","50 Alternating A↔B",
     "50 alternating A→B / B→A transfers; 0 errors",
     "TBUF,RBUF","TBIR,RIR","errors=0","Stress Coverage","P2","Pass"),
    ("STR-003","Stress","10-cycle Overrun Recovery",
     "10 fill→overrun→drain cycles; EIR_B count=10, RIR_B count=10",
     "RBUF,STATUS","RIR,EIR","EIR=10 RIR=10","Stress Coverage","P2","Pass"),

    ("NEG-001","Negative","BG Write No Crash",
     "Write max BG value; verify transfer still works",
     "BG","None","VP continues; transfer correct","Negative Coverage","P1","Pass"),
    ("NEG-002","Negative","FDR Write No Crash",
     "Write FDR; verify VP does not crash",
     "FDR","None","VP continues; transfer correct","Negative Coverage","P1","Pass"),
    ("NEG-003","Negative","STATUS Write-0 Preserves Bits",
     "Write 0 to STATUS; verify bits not cleared",
     "STATUS","TBIR","TBIR preserved after write-0","Register Coverage","P1","Pass"),
    ("NEG-004","Negative","STATUS Write All-1s Clears",
     "Write 0xFFFFFFFF; verify all bits cleared",
     "STATUS","TBIR","STATUS==0 after all-1s write","Register Coverage","P1","Pass"),
    ("NEG-005","Negative","Overrun No EIR Without OEN",
     "Overrun with OEN=0; EIR must be absent",
     "CON,STATUS","EIR","EIR not in log","Error Coverage","P1","Pass"),
    ("NEG-006","Negative","Reserved CON Bits No Side Effects",
     "Write all-1s to CON; transfer still works",
     "CON,TBUF,RBUF","TBIR,RIR","Transfer correct after all-1s CON","Negative Coverage","P2","Pass"),
    ("NEG-007","Negative","Unknown Register Offset",
     "Write to offset 0x30; VP must not crash",
     "None","None","VP continues; transfer unaffected","Negative Coverage","P2","Pass"),

    ("DT-001","Datatype","256-value Full Sweep A→B",
     "All 256 byte values (0x00..0xFF) transferred A→B; 0 errors",
     "TBUF,RBUF","TBIR,RIR","errors=0/256","Data Coverage","P1","Pass"),
    ("DT-002","Datatype","Critical Bit Patterns A→B",
     "0x00,0xFF,0x55,0xAA,0x0F,0xF0 transferred A→B; all correct",
     "TBUF,RBUF","TBIR,RIR","All 6 patterns correct","Data Coverage","P1","Pass"),
    ("DT-003","Datatype","Critical Bit Patterns B→A",
     "Same 6 patterns transferred B→A; all correct",
     "TBUF,RBUF","TBIR,RIR","All 6 patterns correct","Data Coverage","P1","Pass"),

    ("MEM-001","MemoryLeak","200 TBIR Cycles",
     "200 TBUF-write/STATUS-clear cycles; 0 errors (Valgrind target)",
     "TBUF,STATUS","TBIR","errors=0/200","Stress Coverage","P2","Pass"),
    ("MEM-002","MemoryLeak","100 RIR Cycles",
     "100 send+receive+drain cycles; 0 errors",
     "TBUF,RBUF","TBIR,RIR","errors=0/100","Stress Coverage","P2","Pass"),
    ("MEM-003","MemoryLeak","50 EIR Cycles",
     "50 overrun+EIR+drain cycles; EIR count=50",
     "RBUF,STATUS","RIR,EIR","EIR count=50","Stress Coverage","P2","Pass"),
]

TP_HEADERS = ["Test ID","Category","Test Name","Objective","Description",
              "Registers Used","Interrupts/Flags","Expected Result",
              "Coverage Type","Priority","Status"]
TP_WIDTHS  = [10, 14, 30, 45, 45, 20, 18, 40, 22, 8, 10]


def build_test_plan():
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Plan"

    apply_header(ws, TP_HEADERS, TP_WIDTHS)

    for r, row in enumerate(TEST_PLAN, 2):
        # Insert Description (duplicate of Objective for now — extend as needed)
        full_row = list(row[:5]) + [row[4]] + list(row[5:])
        ws.append(full_row[:11])
        fill = ROW_EVEN if r % 2 == 0 else ROW_ODD
        if row[-1] == "Fail":  fill = FAIL_FILL
        elif row[-1] == "Finding": fill = WARN_FILL
        elif row[-1] == "Pass":    fill = PASS_FILL if r % 2 == 0 else ROW_ODD
        style_row(ws, r, fill)

    auto_row_height(ws)
    return wb


# ══════════════════════════════════════════════════════════════════════════════
# 2. usart_errata.xlsx
# ══════════════════════════════════════════════════════════════════════════════
ERRATA = [
    ("BUG-007","RST-002..005","Hardware Reset Pin (RES) Not Implemented",
     "RES input pin must clear all registers and internal state on assertion",
     "Usart2 has no sc_in<bool> rst port and no SWRST register. Offset 0x14 "
     "falls into default: no-op case in b_transport. CON/STATUS/rbuf_full "
     "survive what should be a reset. m_frame_done_ev is not cancelled.",
     "RES pin not modelled in Usart2 sc_module",
     "Option A: add sc_in<bool> rst port; drive from sc_main. "
     "Option B: add OFF_SWRST=0x14 case — on write 1: m_con=0, m_status=0, "
     "m_rbuf=0, m_rbuf_full=false, m_frame_done_ev.cancel(), sync_irq_trace()",
     "High","Open","VP Team"),
    ("BUG-001","RW-002","CON Reserved Bits [31:16] Stored",
     "CON[31:16] must read 0 (reserved per spec)",
     "VP model stores full 32-bit m_con including reserved bits",
     "Model stores raw register write without masking",
     "Mask reserved bits: m_con = value & 0x0000FFFFu in b_transport",
     "Low","Open","VP Team"),
    ("BUG-002","RW-006","BG Register Not Implemented",
     "BG at offset 0x0C: 13-bit reload register + live downcounter",
     "VP model b_transport default case — BG is a no-op; reads return 0",
     "Registers beyond STATUS (0x20) not implemented in Usart2::b_transport",
     "Add OFF_BG case; implement m_bg_reload and simulated downcounter",
     "Medium","Open","VP Team"),
    ("BUG-003","RW-007","FDR Register Not Implemented",
     "FDR at offset 0x10: STEP[7:0]+DM[8]; active when CON.FDE=1",
     "VP model b_transport default case — FDR is a no-op",
     "Registers beyond STATUS (0x20) not implemented",
     "Add OFF_FDR case; implement sigma-delta accumulator logic",
     "Medium","Open","VP Team"),
    ("BUG-004","LB-004","CON.LB Loopback Not Functional",
     "CON.LB=1 must internally connect TSR output to RSR input",
     "CON.LB is stored but sc_fifo routing is hard-wired in sc_main; loopback has no effect",
     "sc_fifo connections are fixed at elaboration time in sc_main",
     "In tx_deliver_method: if (m_con & CON_LB) write to own rx_port else tx_port",
     "Medium","Open","VP Team"),
    ("BUG-005","ADV-006","RBUF Stale Value After rbuf_full Clear",
     "Spec: RBUF holds valid data until rbuf_full asserted; behaviour after clear is impl-defined",
     "VP model m_rbuf retains last byte; second read after drain returns stale value",
     "m_rbuf is not cleared on RBUF read; only m_rbuf_full is cleared",
     "Document VP behaviour explicitly; add note in usart2.h",
     "Info","Documented","VP Team"),
    ("BUG-006","FN-003/ADV-001","EIR Gating Depends on CON_OEN",
     "EIR must fire simultaneously with RIR only when CON.OEN=1",
     "VP model checks (m_con & CON_OEN) in rx_thread — correct behaviour",
     "N/A — model implements this correctly",
     "No fix required — compliant",
     "None","Closed","N/A"),
]

ERR_HEADERS = ["Bug ID","Test ID","Failure Summary","Expected","Actual",
               "Root Cause","Suggested Fix","Severity","Status","Owner"]
ERR_WIDTHS  = [10, 10, 35, 40, 40, 45, 50, 10, 10, 12]


def build_errata():
    wb = Workbook()
    ws = wb.active
    ws.title = "Errata"

    apply_header(ws, ERR_HEADERS, ERR_WIDTHS)

    sev_fill = {
        "Low":    PatternFill("solid", fgColor="FFEB9C"),
        "Medium": PatternFill("solid", fgColor="FFC7CE"),
        "Info":   PatternFill("solid", fgColor="C6EFCE"),
        "None":   PatternFill("solid", fgColor="C6EFCE"),
    }

    for r, row in enumerate(ERRATA, 2):
        ws.append(list(row))
        fill = sev_fill.get(row[7], ROW_ODD)
        style_row(ws, r, fill)

    auto_row_height(ws)
    return wb


# ══════════════════════════════════════════════════════════════════════════════
# 3. coverage_summary.xlsx
# ══════════════════════════════════════════════════════════════════════════════
COV_DATA = [
    # (Category, Tests, Planned, Passed, Findings, Coverage%)
    # Actual run results — 20260515
    ("Reset Coverage",      5,  5,  1, 0,  20),   # RST-001 PASS; RST-002..005 FAIL (BUG-007)
    ("Register Coverage",   7,  7,  7, 3, 100),   # RW-001..007 all pass (findings: BUG-001/002/003)
    ("Functional Coverage", 5,  5,  5, 0, 100),
    ("Interrupt Coverage",  4,  4,  4, 0, 100),
    ("Error Coverage",      6,  6,  6, 0, 100),
    ("Timing Coverage",     4,  4,  4, 0, 100),
    ("Data Coverage",       5,  5,  5, 1, 100),   # LB-004 finding (CON.LB no-op)
    ("Stress Coverage",     5,  5,  5, 0, 100),
    ("Performance Coverage",4,  4,  4, 0, 100),
    ("Negative Coverage",   7,  7,  7, 0, 100),
    ("Memory Stability",    3,  3,  3, 0, 100),
]

COV_HEADERS = ["Coverage Type","Total Tests","Planned","Passed","Findings","Coverage %"]
COV_WIDTHS  = [28, 14, 12, 12, 12, 14]


def build_coverage():
    wb = Workbook()
    ws = wb.active
    ws.title = "Coverage Summary"

    apply_header(ws, COV_HEADERS, COV_WIDTHS)

    for r, row in enumerate(COV_DATA, 2):
        ws.append(list(row))
        pct = row[5]
        if pct == 100:   fill = PASS_FILL
        elif pct >= 90:  fill = PatternFill("solid", fgColor="FFEB9C")
        else:            fill = FAIL_FILL
        style_row(ws, r, fill)

    # Totals row
    r_tot = len(COV_DATA) + 2
    totals = ["TOTAL",
              sum(x[1] for x in COV_DATA),
              sum(x[2] for x in COV_DATA),
              sum(x[3] for x in COV_DATA),
              sum(x[4] for x in COV_DATA),
              round(sum(x[5] for x in COV_DATA) / len(COV_DATA))]
    ws.append(totals)
    for cell in ws[r_tot]:
        cell.font      = Font(name="Calibri", size=11, bold=True)
        cell.fill      = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = CENTER
        cell.border    = BORDER

    auto_row_height(ws)
    return wb


# ── Entry point ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    tp = build_test_plan()
    tp.save(os.path.join(DOCS_DIR, "USART_TestPlan.xlsx"))
    print(f"  Written: {os.path.join(DOCS_DIR, 'USART_TestPlan.xlsx')}")

    er = build_errata()
    er.save(os.path.join(DOCS_DIR, "usart_errata.xlsx"))
    print(f"  Written: {os.path.join(DOCS_DIR, 'usart_errata.xlsx')}")

    cv = build_coverage()
    cv.save(os.path.join(DOCS_DIR, "coverage_summary.xlsx"))
    print(f"  Written: {os.path.join(DOCS_DIR, 'coverage_summary.xlsx')}")

    print("\nAll XLSX deliverables generated successfully.")
